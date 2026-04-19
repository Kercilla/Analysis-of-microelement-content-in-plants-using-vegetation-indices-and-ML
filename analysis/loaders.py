import re
import subprocess
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import geopandas as gpd

def load_multi(
    gpkg_pattern: str,
    band_map: dict[str, str],
    value_col: str = "_mean",
    id_col: str = "id",
) -> pd.DataFrame:
    
    from glob import glob
    files = sorted(glob(gpkg_pattern))
    if not files:
        raise FileNotFoundError(f"Нет файлов по паттерну: {gpkg_pattern}")

    records = {}
    for fpath in files:
        stem = Path(fpath).stem.split("_")[0]  # "11_05_29" → "11"
        channel = band_map.get(stem)
        if channel is None:
            continue

        gdf = gpd.read_file(fpath)
        for _, row in gdf.iterrows():
            pid = int(row[id_col])
            val = row[value_col]
            if pd.notna(val):
                if pid not in records:
                    records[pid] = {}
                records[pid][channel] = float(val)

    df = pd.DataFrame.from_dict(records, orient="index").sort_index()
    df.index.name = "id"
    return df

def _parse_hyper_channel(filename: str, prefix_len: int = 1) -> int | None:
    
    stem = Path(filename).stem
    if not stem.isdigit():
        return None
    channel_str = stem[prefix_len:]
    if not channel_str:
        return None
    return int(channel_str)

def load_hyper_wavelength_map(xlsx_path: str) -> dict[int, float]:
    
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    wl_row = [cell.value for cell in ws[1]]
    wl_map = {}
    for i, val in enumerate(wl_row[5:], start=1):
        if val is not None and isinstance(val, (int, float)):
            wl_map[i] = float(val)
    return wl_map

def load_hyper_date(
    date_folder: str,
    wl_map: dict[int, float],
    prefix_len: int = 1,
    value_col: str = "_mean",
    id_col: str = "id",
    min_valid_bands: int = 50,
) -> tuple[pd.DataFrame, np.ndarray]:
    
    date_folder = Path(date_folder)
    if not date_folder.exists():
        raise FileNotFoundError(f"Папка не найдена: {date_folder}")

    subfolders = sorted([d for d in date_folder.iterdir() if d.is_dir()])
    print(f"  Папка: {date_folder.name}, наборов: {len(subfolders)}")

    if not subfolders:
        raise ValueError(f"Нет подпапок в {date_folder}")

    all_values = defaultdict(lambda: defaultdict(list))
    sets_with_data = 0

    for si, subfolder in enumerate(subfolders):
        gpkg_files = list(subfolder.glob("*.gpkg"))
        if not gpkg_files:
            continue

        set_has_data = False
        for fpath in gpkg_files:
            ch = _parse_hyper_channel(fpath.name, prefix_len)
            if ch is None or ch not in wl_map:
                continue

            try:
                gdf = gpd.read_file(fpath)
            except Exception:
                continue

            if value_col not in gdf.columns:
                continue

            for _, row in gdf.iterrows():
                pid = int(row[id_col])
                val = row[value_col]
                if pd.notna(val) and val != 0:
                    all_values[pid][ch].append(float(val))
                    set_has_data = True

        if set_has_data:
            sets_with_data += 1

        if (si + 1) % 20 == 0 or si == len(subfolders) - 1:
            print(f"    Прочитано {si+1}/{len(subfolders)} наборов "
                  f"({sets_with_data} с данными)")

    sorted_channels = sorted(wl_map.keys())
    wavelengths = np.array([wl_map[ch] for ch in sorted_channels])
    col_names = [f"{wl_map[ch]:.2f}" for ch in sorted_channels]

    rows = {}
    for pid, channels in all_values.items():
        spectrum = []
        valid_count = 0
        for ch in sorted_channels:
            values = channels.get(ch, [])
            if values:
                spectrum.append(np.mean(values))
                valid_count += 1
            else:
                spectrum.append(np.nan)

        if valid_count >= min_valid_bands:
            rows[pid] = spectrum

    df = pd.DataFrame.from_dict(rows, orient="index", columns=col_names)
    df.index.name = "id"
    df = df.sort_index()

    n_sets_per_point = {
        pid: np.mean([len(channels.get(ch, [])) for ch in sorted_channels if channels.get(ch)])
        for pid, channels in all_values.items()
        if pid in rows
    }
    avg_sets = np.mean(list(n_sets_per_point.values())) if n_sets_per_point else 0

    print(f"  Результат: {len(df)} точек, {len(wavelengths)} каналов, "
          f"среднее {avg_sets:.1f} наборов/точку")

    return df, wavelengths

def load_hyper_from_xlsx(xlsx_path: str) -> tuple[pd.DataFrame, np.ndarray]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    wl_row = [cell.value for cell in ws[1]]

    # Auto-detect: find first column where numeric values in range 300-2500 nm start
    def _to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).strip())
        except (ValueError, TypeError):
            return None

    wl_start = None
    for i, v in enumerate(wl_row):
        fv = _to_float(v)
        if fv is not None and 300 <= fv <= 2500:
            wl_start = i
            break

    if wl_start is None:
        for i, v in enumerate(wl_row):
            if _to_float(v) is not None:
                wl_start = i
                break

    if wl_start is None:
        raise ValueError(
            f"load_hyper_from_xlsx: не найдены длины волн в первой строке файла {xlsx_path}. "
            f"Первые 10 значений заголовка: {wl_row[:10]}"
        )

    wavelengths = np.array([
        fv for v in wl_row[wl_start:]
        if (fv := _to_float(v)) is not None
    ])

    # id column: search for first column containing sequential integers (sample IDs)
    # Try column index 1 first (standard), then scan
    id_col = 1
    for row in ws.iter_rows(min_row=3, max_row=min(5, ws.max_row), values_only=True):
        if row[1] is not None and isinstance(row[1], (int, float)):
            id_col = 1
            break
        if row[0] is not None and isinstance(row[0], (int, float)):
            id_col = 0
            break

    records = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[id_col] is None:
            continue
        pid_f = _to_float(row[id_col])
        if pid_f is None:
            continue
        pid = int(pid_f)
        # Skip stdev rows: check cell just before wl_start
        if wl_start > 0 and str(row[wl_start - 1]).strip().lower() == "stdev":
            continue
        spectrum = [
            fv if (fv := _to_float(v)) is not None else np.nan
            for v in row[wl_start:wl_start + len(wavelengths)]
        ]
        if pid not in records:
            records[pid] = []
        records[pid].append(spectrum)

    data = {pid: np.nanmean(specs, axis=0) for pid, specs in records.items()}
    col_names = [f"{w:.2f}" for w in wavelengths]
    df = pd.DataFrame.from_dict(data, orient="index", columns=col_names)
    df.index.name = "id"
    df = df.sort_index()

    print(f"  Гипер xlsx: {len(df)} точек, {len(wavelengths)} каналов "
          f"({wavelengths[0]:.0f}–{wavelengths[-1]:.0f} нм)")
    return df, wavelengths

def load_chemistry_doc(
    filepath: str,
    columns: list[str],
    id_offset: int = 0,
) -> pd.DataFrame:
    
    result = subprocess.run(
        ["antiword", "-m", "UTF-8.txt", filepath],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"antiword error: {result.stderr}")

    parts = result.stdout.split("|")
    values = []
    for p in parts:
        p = p.strip().replace(",", ".")
        if p and re.match(r"^[\d\.<]+", p):
            values.append(p)

    samples = []
    i = 0
    while i < len(values) - 16:
        try:
            sid = int(values[i])
            if 1 <= sid <= 9999:
                row_vals = values[i + 1 : i + 17]
                parsed = [sid - id_offset]
                for v in row_vals:
                    parsed.append(np.nan if v.startswith("<") else float(v))
                samples.append(parsed)
                i += 17
                continue
        except (ValueError, IndexError):
            pass
        i += 1

    df = pd.DataFrame(samples, columns=["id"] + columns)
    return df.set_index("id").sort_index()

def load_chemistry(filepath: str, columns: list[str], id_offset: int = 0) -> pd.DataFrame:
    
    ext = Path(filepath).suffix.lower()
    if ext == ".doc":
        return load_chemistry_doc(filepath, columns, id_offset)
    elif ext in (".xlsx", ".xls"):
        df = pd.read_excel(filepath, index_col=0)
        if id_offset:
            df.index = df.index - id_offset
        return df.sort_index()
    elif ext == ".csv":
        return pd.read_csv(filepath, index_col=0).sort_index()
    raise ValueError(f"Неподдерживаемый формат: {ext}")


def load_hyper_from_csv(csv_path: str) -> tuple[pd.DataFrame, np.ndarray]:
    df = pd.read_csv(csv_path, index_col=0)
    df.index.name = "id"
    wavelengths = np.array([float(c) for c in df.columns])
    print(f"  Hyper CSV: {len(df)} pts, {len(wavelengths)} bands "
          f"({wavelengths[0]:.0f}-{wavelengths[-1]:.0f} nm)")
    return df, wavelengths


def load_hyper_auto(cfg, date_key):
    paths = cfg["paths"]["hyper"]
    hn = cfg["hyper_naming"]

    for key, loader in [
        (f"{date_key}_csv", lambda p: load_hyper_from_csv(p)),
        (f"{date_key}_xlsx", lambda p: load_hyper_from_xlsx(p)),
    ]:
        fpath = paths.get(key)
        if fpath and Path(fpath).exists():
            return loader(fpath)

    folder = paths.get(f"{date_key}_folder")
    if folder and Path(folder).exists():
        wl_map = load_hyper_wavelength_map(paths["wavelength_map"])
        return load_hyper_date(folder, wl_map, hn["prefix_length"],
                               hn["value_column"], hn["id_column"],
                               hn["min_valid_bands"])

    raise FileNotFoundError(f"No hyper data for {date_key}")